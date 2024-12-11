import sys
import sqlite3
import bcrypt
from PyQt6.QtWidgets import QApplication, QMainWindow, QVBoxLayout, QWidget, QLabel, QLineEdit, QPushButton, QTableWidget, QTableWidgetItem, QHeaderView, QMessageBox, QComboBox, QInputDialog, QTabWidget, QGridLayout, QFrame, QTextEdit, QDialog, QFormLayout
from PyQt6.QtCore import Qt
from PyQt6.QtGui import QFont, QColor
import openpyxl
import re
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# Подключение к базе данных
conn = sqlite3.connect('Dantistikss.db')
cursor = conn.cursor()

# Функция для получения данных из таблицы Patient
def get_patients():
    cursor.execute("SELECT * FROM Patient")
    return cursor.fetchall()

# Функция для получения данных из таблицы Doctor
def get_doctors():
    cursor.execute("SELECT * FROM Doctor")
    return cursor.fetchall()

# Функция для получения данных из таблицы Appointment
def get_appointments():
    cursor.execute("""
        SELECT a.appointment_id, p.first_name || ' ' || p.middle_name || ' ' || p.last_name AS patient_name, 
               d.first_name || ' ' || d.middle_name || ' ' || d.last_name AS doctor_name, a.appointment_date, a.status, d.specialization
        FROM Appointment a
        JOIN Patient p ON a.patient_id = p.patient_id
        JOIN Doctor d ON a.doctor_id = d.doctor_id
    """)
    return cursor.fetchall()

# Функция для отображения данных в таблице
def show_patients(table):
    patients = get_patients()
    table.setRowCount(len(patients))
    for row, patient in enumerate(patients):
        for col, item in enumerate(patient):
            table.setItem(row, col, QTableWidgetItem(str(item)))

# Функция для отображения данных о врачах в таблице
def show_doctors(table):
    doctors = get_doctors()
    table.setRowCount(len(doctors))
    for row, doctor in enumerate(doctors):
        for col, item in enumerate(doctor):
            table.setItem(row, col, QTableWidgetItem(str(item)))

# Функция для отображения данных о приёмах в таблице
def show_appointments(table):
    appointments = get_appointments()
    table.setRowCount(len(appointments))
    for row, appointment in enumerate(appointments):
        for col, item in enumerate(appointment):
            table.setItem(row, col, QTableWidgetItem(str(item)))

# Окно авторизации
class LoginWindow(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Авторизация")
        self.setGeometry(100, 100, 300, 200)

        layout = QVBoxLayout()

        self.username_input = QLineEdit(self)
        self.username_input.setPlaceholderText("Имя пользователя")
        layout.addWidget(self.username_input)

        self.password_input = QLineEdit(self)
        self.password_input.setPlaceholderText("Пароль")
        self.password_input.setEchoMode(QLineEdit.EchoMode.Password)
        layout.addWidget(self.password_input)

        self.login_button = QPushButton("Войти", self)
        self.login_button.clicked.connect(self.login)
        layout.addWidget(self.login_button)

        self.register_button = QPushButton("Регистрация пациента", self)
        self.register_button.clicked.connect(self.show_patient_register_window)
        layout.addWidget(self.register_button)

        self.setLayout(layout)

    def login(self):
        username = self.username_input.text()
        password = self.password_input.text().encode('utf-8')

        # Проверка учетных записей администраторов
        cursor.execute("SELECT * FROM Administrator WHERE username=?", (username,))
        admin = cursor.fetchone()

        if admin:
            if bcrypt.checkpw(password, admin[2].encode('utf-8')):
                self.main_window = MainWindow(admin[3])  # Передаем роль пользователя
                self.main_window.show()
                self.close()
                return

        # Проверка учетных записей пациентов
        cursor.execute("SELECT * FROM Patient WHERE username=?", (username,))
        patient = cursor.fetchone()

        if patient:
            if bcrypt.checkpw(password, patient[13].encode('utf-8')):
                self.patient_window = PatientWindow(patient)
                self.patient_window.show()
                self.close()
                return

        # Проверка учетных записей врачей
        cursor.execute("SELECT * FROM Doctor WHERE username=?", (username,))
        doctor = cursor.fetchone()

        if doctor:
            if bcrypt.checkpw(password, doctor[7].encode('utf-8')):
                self.doctor_window = DoctorWindow(doctor)
                self.doctor_window.show()
                self.close()
                return

        QMessageBox.warning(self, "Ошибка", "Неверное имя пользователя или пароль")

    def show_patient_register_window(self):
        self.patient_register_window = PatientRegisterWindow()
        self.patient_register_window.show()

# Класс Validator
class Validator:
    @staticmethod
    def validate_name(name):
        return re.match(r'^[А-ЯЁ][а-яё]+$', name) is not None

    @staticmethod
    def validate_gender(gender):
        return gender in ['М', 'Ж']

    @staticmethod
    def validate_phone(phone):
        return re.match(r'^\+7\d{10}$', phone) is not None

    @staticmethod
    def validate_date(date):
        return re.match(r'^\d{4}-\d{2}-\d{2}$', date) is not None

    @staticmethod
    def validate_passport_series(series):
        return re.match(r'^\d{4}$', series) is not None

    @staticmethod
    def validate_passport_number(number):
        return re.match(r'^\d{6}$', number) is not None

    @staticmethod
    def validate_insurance_number(number):
        return re.match(r'^\d{16}$', number) is not None

    @staticmethod
    def validate_address(address):
        return re.match(r'^[А-ЯЁа-яё0-9\s,-]+$', address) is not None

# Окно регистрации пациента
class PatientRegisterWindow(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Регистрация пациента")
        self.setGeometry(100, 100, 300, 500)

        layout = QVBoxLayout()

        self.username_input = QLineEdit(self)
        self.username_input.setPlaceholderText("Имя пользователя")
        layout.addWidget(self.username_input)

        self.password_input = QLineEdit(self)
        self.password_input.setPlaceholderText("Пароль")
        self.password_input.setEchoMode(QLineEdit.EchoMode.Password)
        layout.addWidget(self.password_input)

        self.confirm_password_input = QLineEdit(self)
        self.confirm_password_input.setPlaceholderText("Подтвердите пароль")
        self.confirm_password_input.setEchoMode(QLineEdit.EchoMode.Password)
        layout.addWidget(self.confirm_password_input)

        self.first_name_input = QLineEdit(self)
        self.first_name_input.setPlaceholderText("Имя")
        layout.addWidget(self.first_name_input)

        self.last_name_input = QLineEdit(self)
        self.last_name_input.setPlaceholderText("Фамилия")
        layout.addWidget(self.last_name_input)

        self.middle_name_input = QLineEdit(self)
        self.middle_name_input.setPlaceholderText("Отчество")
        layout.addWidget(self.middle_name_input)

        self.birth_date_input = QLineEdit(self)
        self.birth_date_input.setPlaceholderText("Дата рождения (YYYY-MM-DD)")
        layout.addWidget(self.birth_date_input)

        self.gender_input = QComboBox(self)
        self.gender_input.addItems(["М", "Ж"])
        layout.addWidget(self.gender_input)

        self.phone_input = QLineEdit(self)
        self.phone_input.setPlaceholderText("Телефон (+7XXXXXXXXXX)")
        layout.addWidget(self.phone_input)

        self.email_input = QLineEdit(self)
        self.email_input.setPlaceholderText("Email")
        layout.addWidget(self.email_input)

        self.address_input = QLineEdit(self)
        self.address_input.setPlaceholderText("Адрес")
        layout.addWidget(self.address_input)

        self.passport_series_input = QLineEdit(self)
        self.passport_series_input.setPlaceholderText("Серия паспорта (4 цифры)")
        layout.addWidget(self.passport_series_input)

        self.passport_number_input = QLineEdit(self)
        self.passport_number_input.setPlaceholderText("Номер паспорта (6 цифр)")
        layout.addWidget(self.passport_number_input)

        self.insurance_number_input = QLineEdit(self)
        self.insurance_number_input.setPlaceholderText("Номер страховки (16 цифр)")
        layout.addWidget(self.insurance_number_input)

        self.register_button = QPushButton("Зарегистрироваться", self)
        self.register_button.clicked.connect(self.register)
        layout.addWidget(self.register_button)

        self.setLayout(layout)

    def register(self):
        username = self.username_input.text()
        password = self.password_input.text().encode('utf-8')
        confirm_password = self.confirm_password_input.text().encode('utf-8')

        if password != confirm_password:
            QMessageBox.warning(self, "Ошибка", "Пароли не совпадают")
            return

        hashed_password = bcrypt.hashpw(password, bcrypt.gensalt())

        first_name = self.first_name_input.text()
        last_name = self.last_name_input.text()
        middle_name = self.middle_name_input.text()
        birth_date = self.birth_date_input.text()
        gender = self.gender_input.currentText()
        phone = self.phone_input.text()
        email = self.email_input.text()
        address = self.address_input.text()
        passport_series = self.passport_series_input.text()
        passport_number = self.passport_number_input.text()
        insurance_number = self.insurance_number_input.text()

        if not Validator.validate_name(first_name):
            QMessageBox.warning(self, "Ошибка", "Имя должно начинаться с заглавной буквы и содержать только русские буквы")
            return

        if not Validator.validate_name(last_name):
            QMessageBox.warning(self, "Ошибка", "Фамилия должна начинаться с заглавной буквы и содержать только русские буквы")
            return

        if not Validator.validate_name(middle_name):
            QMessageBox.warning(self, "Ошибка", "Отчество должно начинаться с заглавной буквы и содержать только русские буквы")
            return

        if not Validator.validate_date(birth_date):
            QMessageBox.warning(self, "Ошибка", "Дата рождения должна быть в формате YYYY-MM-DD")
            return

        if not Validator.validate_gender(gender):
            QMessageBox.warning(self, "Ошибка", "Пол должен быть 'М' или 'Ж'")
            return

        if not Validator.validate_phone(phone):
            QMessageBox.warning(self, "Ошибка", "Телефон должен быть в формате +7XXXXXXXXXX")
            return

        if not Validator.validate_address(address):
            QMessageBox.warning(self, "Ошибка", "Адрес должен содержать только русские буквы, цифры и знаки препинания")
            return

        if not Validator.validate_passport_series(passport_series):
            QMessageBox.warning(self, "Ошибка", "Серия паспорта должна состоять из 4 цифр")
            return

        if not Validator.validate_passport_number(passport_number):
            QMessageBox.warning(self, "Ошибка", "Номер паспорта должен состоять из 6 цифр")
            return

        if not Validator.validate_insurance_number(insurance_number):
            QMessageBox.warning(self, "Ошибка", "Номер страховки должен состоять из 16 цифр")
            return

        try:
            cursor.execute("""
                INSERT INTO Patient (first_name, last_name, middle_name, birth_date, gender,
                 phone, email, address, passport_series, passport_number, insurance_number, username, password)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (first_name, last_name, middle_name, birth_date, gender, phone, email,
                  address, passport_series, passport_number, insurance_number, username, hashed_password.decode('utf-8')))
            conn.commit()
            QMessageBox.information(self, "Успех", "Регистрация прошла успешно")
            self.close()
        except sqlite3.Error as e:
            QMessageBox.warning(self, "Ошибка", f"Ошибка при регистрации: {e}")

# Окно регистрации врача (только для админа)
class DoctorRegisterWindow(QWidget):
    def __init__(self, main_window):
        super().__init__()
        self.main_window = main_window
        self.setWindowTitle("Регистрация врача")
        self.setGeometry(100, 100, 300, 400)

        layout = QVBoxLayout()

        self.username_input = QLineEdit(self)
        self.username_input.setPlaceholderText("Имя пользователя")
        layout.addWidget(self.username_input)

        self.password_input = QLineEdit(self)
        self.password_input.setPlaceholderText("Пароль")
        self.password_input.setEchoMode(QLineEdit.EchoMode.Password)
        layout.addWidget(self.password_input)

        self.confirm_password_input = QLineEdit(self)
        self.confirm_password_input.setPlaceholderText("Подтвердите пароль")
        self.confirm_password_input.setEchoMode(QLineEdit.EchoMode.Password)
        layout.addWidget(self.confirm_password_input)

        self.first_name_input = QLineEdit(self)
        self.first_name_input.setPlaceholderText("Имя")
        layout.addWidget(self.first_name_input)

        self.last_name_input = QLineEdit(self)
        self.last_name_input.setPlaceholderText("Фамилия")
        layout.addWidget(self.last_name_input)

        self.middle_name_input = QLineEdit(self)
        self.middle_name_input.setPlaceholderText("Отчество")
        layout.addWidget(self.middle_name_input)

        self.specialization_input = QLineEdit(self)
        self.specialization_input.setPlaceholderText("Специализация")
        layout.addWidget(self.specialization_input)

        self.phone_input = QLineEdit(self)
        self.phone_input.setPlaceholderText("Телефон")
        layout.addWidget(self.phone_input)

        self.email_input = QLineEdit(self)
        self.email_input.setPlaceholderText("Email")
        layout.addWidget(self.email_input)

        self.register_button = QPushButton("Зарегистрироваться", self)
        self.register_button.clicked.connect(self.register)
        layout.addWidget(self.register_button)

        self.setLayout(layout)

    def register(self):
        username = self.username_input.text()
        password = self.password_input.text().encode('utf-8')
        confirm_password = self.confirm_password_input.text().encode('utf-8')

        if password != confirm_password:
            QMessageBox.warning(self, "Ошибка", "Пароли не совпадают")
            return

        hashed_password = bcrypt.hashpw(password, bcrypt.gensalt())

        first_name = self.first_name_input.text()
        last_name = self.last_name_input.text()
        middle_name = self.middle_name_input.text()
        specialization = self.specialization_input.text()
        phone = self.phone_input.text()
        email = self.email_input.text()

        try:
            cursor.execute("""
                INSERT INTO Doctor (first_name, last_name, middle_name, specialization, phone, email, password, username)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """, (first_name, last_name, middle_name, specialization, phone, email, hashed_password.decode('utf-8'),
                  username))
            conn.commit()
            QMessageBox.information(self, "Успех", "Регистрация прошла успешно")
            self.main_window.update_doctor_table()  # Обновляем таблицу врачей
            self.close()
        except sqlite3.Error as e:
            QMessageBox.warning(self, "Ошибка", f"Ошибка при регистрации: {e}")

# Окно добавления пациента (только для админа)
class AddPatientWindow(QWidget):
    def __init__(self, main_window):
        super().__init__()
        self.main_window = main_window
        self.setWindowTitle("Добавление пациента")
        self.setGeometry(100, 100, 300, 350)  # Увеличиваем размер окна для новых полей

        layout = QVBoxLayout()

        self.username_input = QLineEdit(self)
        self.username_input.setPlaceholderText("Логин пациента")
        layout.addWidget(self.username_input)

        self.password_input = QLineEdit(self)
        self.password_input.setPlaceholderText("Пароль пациента")
        self.password_input.setEchoMode(QLineEdit.EchoMode.Password)  # Скрываем пароль
        layout.addWidget(self.password_input)

        self.first_name_input = QLineEdit(self)
        self.first_name_input.setPlaceholderText("Имя")
        layout.addWidget(self.first_name_input)

        self.last_name_input = QLineEdit(self)
        self.last_name_input.setPlaceholderText("Фамилия")
        layout.addWidget(self.last_name_input)

        self.middle_name_input = QLineEdit(self)
        self.middle_name_input.setPlaceholderText("Отчество")
        layout.addWidget(self.middle_name_input)

        self.birth_date_input = QLineEdit(self)
        self.birth_date_input.setPlaceholderText("Дата рождения (YYYY-MM-DD) ")
        layout.addWidget(self.birth_date_input)

        self.gender_input = QComboBox(self)
        self.gender_input.addItems(["М", "Ж"])
        layout.addWidget(self.gender_input)

        self.phone_input = QLineEdit(self)
        self.phone_input.setPlaceholderText("Телефон (+7XXXXXXXXXX)")
        layout.addWidget(self.phone_input)

        self.email_input = QLineEdit(self)
        self.email_input.setPlaceholderText("Email")
        layout.addWidget(self.email_input)

        self.address_input = QLineEdit(self)
        self.address_input.setPlaceholderText("Адрес")
        layout.addWidget(self.address_input)

        self.passport_series_input = QLineEdit(self)
        self.passport_series_input.setPlaceholderText("Серия паспорта (4 цифры)")
        layout.addWidget(self.passport_series_input)

        self.passport_number_input = QLineEdit(self)
        self.passport_number_input.setPlaceholderText("Номер паспорта (6 цифр)")
        layout.addWidget(self.passport_number_input)

        self.insurance_number_input = QLineEdit(self)
        self.insurance_number_input.setPlaceholderText("Номер страховки (16 цифр)")
        layout.addWidget(self.insurance_number_input)

        self.add_button = QPushButton("Добавить", self)
        self.add_button.clicked.connect(self.add_patient)
        layout.addWidget(self.add_button)

        self.setLayout(layout)

    def add_patient(self):
        username = self.username_input.text()
        password = self.password_input.text()

        # Проверка уникальности логина пациента
        if self.validate_username(username):
            QMessageBox.warning(self, "Ошибка", "Логин пациента уже существует")
            return

        first_name = self.first_name_input.text()
        last_name = self.last_name_input.text()
        middle_name = self.middle_name_input.text()
        birth_date = self.birth_date_input.text()
        gender = self.gender_input.currentText()
        phone = self.phone_input.text()
        email = self.email_input.text()
        address = self.address_input.text()
        passport_series = self.passport_series_input.text()
        passport_number = self.passport_number_input.text()
        insurance_number = self.insurance_number_input.text()

        if not Validator.validate_name(first_name):
            QMessageBox.warning(self, "Ошибка", "Имя должно начинаться с заглавной буквы и содержать только русские буквы")
            return

        if not Validator.validate_name(last_name):
            QMessageBox.warning(self, "Ошибка", "Фамилия должна начинаться с заглавной буквы и содержать только русские буквы")
            return

        if not Validator.validate_name(middle_name):
            QMessageBox.warning(self, "Ошибка", "Отчество должно начинаться с заглавной буквы и содержать только русские буквы")
            return

        if not Validator.validate_date(birth_date):
            QMessageBox.warning(self, "Ошибка", "Дата рождения должна быть в формате YYYY-MM-DD")
            return

        if not Validator.validate_gender(gender):
            QMessageBox.warning(self, "Ошибка", "Пол должен быть 'М' или 'Ж'")
            return

        if not Validator.validate_phone(phone):
            QMessageBox.warning(self, "Ошибка", "Телефон должен быть в формате +7XXXXXXXXXX")
            return

        if not Validator.validate_address(address):
            QMessageBox.warning(self, "Ошибка", "Адрес должен содержать только русские буквы, цифры и знаки препинания")
            return

        if not Validator.validate_passport_series(passport_series):
            QMessageBox.warning(self, "Ошибка", "Серия паспорта должна состоять из 4 цифр")
            return

        if not Validator.validate_passport_number(passport_number):
            QMessageBox.warning(self, "Ошибка", "Номер паспорта должен состоять из 6 цифр")
            return

        if not Validator.validate_insurance_number(insurance_number):
            QMessageBox.warning(self, "Ошибка", "Номер страховки должен состоять из 16 цифр")
            return

        try:
            cursor.execute("""
                INSERT INTO Patient (username, password, first_name, last_name, middle_name,
                 birth_date, gender, phone, email, address, passport_series, passport_number, insurance_number)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (username, password, first_name, last_name, middle_name,
                  birth_date, gender, phone, email, address, passport_series, passport_number, insurance_number))
            conn.commit()
            QMessageBox.information(self, "Успех", "Пациент успешно добавлен")
            self.main_window.update_patient_table()  # Обновляем таблицу пациентов
            self.close()
        except sqlite3.Error as e:
            QMessageBox.warning(self, "Ошибка", f"Ошибка при добавлении пациента: {e}")

    def validate_username(self, username):
        # Проверка уникальности логина пациента в базе данных
        cursor.execute("SELECT * FROM Patient WHERE username = ?", (username,))
        return cursor.fetchone() is not None

# Окно редактирования данных пациента (только для админа)
class EditPatientWindow(QWidget):
    def __init__(self, main_window, patient_id):
        super().__init__()
        self.main_window = main_window
        self.patient_id = patient_id
        self.setWindowTitle("Редактирование данных пациента")
        self.setGeometry(100, 100, 300, 250)

        layout = QVBoxLayout()

        self.first_name_input = QLineEdit(self)
        self.first_name_input.setPlaceholderText("Имя")
        layout.addWidget(self.first_name_input)

        self.last_name_input = QLineEdit(self)
        self.last_name_input.setPlaceholderText("Фамилия")
        layout.addWidget(self.last_name_input)

        self.middle_name_input = QLineEdit(self)
        self.middle_name_input.setPlaceholderText("Отчество")
        layout.addWidget(self.middle_name_input)

        self.birth_date_input = QLineEdit(self)
        self.birth_date_input.setPlaceholderText("Дата рождения")
        layout.addWidget(self.birth_date_input)

        self.gender_input = QComboBox(self)
        self.gender_input.addItems(["М", "Ж"])
        layout.addWidget(self.gender_input)

        self.phone_input = QLineEdit(self)
        self.phone_input.setPlaceholderText("Телефон (+7XXXXXXXXXX)")
        layout.addWidget(self.phone_input)

        self.email_input = QLineEdit(self)
        self.email_input.setPlaceholderText("Email")
        layout.addWidget(self.email_input)

        self.address_input = QLineEdit(self)
        self.address_input.setPlaceholderText("Адрес")
        layout.addWidget(self.address_input)

        self.passport_series_input = QLineEdit(self)
        self.passport_series_input.setPlaceholderText("Серия паспорта (4 цифры)")
        layout.addWidget(self.passport_series_input)

        self.passport_number_input = QLineEdit(self)
        self.passport_number_input.setPlaceholderText("Номер паспорта (6 цифр)")
        layout.addWidget(self.passport_number_input)

        self.insurance_number_input = QLineEdit(self)
        self.insurance_number_input.setPlaceholderText("Номер страховки (16 цифр)")
        layout.addWidget(self.insurance_number_input)

        self.save_button = QPushButton("Сохранить", self)
        self.save_button.clicked.connect(self.save_patient)
        layout.addWidget(self.save_button)

        self.setLayout(layout)

        self.load_patient_data()

    def load_patient_data(self):
        cursor.execute("SELECT * FROM Patient WHERE patient_id=?", (self.patient_id,))
        patient = cursor.fetchone()
        if patient:
            self.first_name_input.setText(patient[1])
            self.last_name_input.setText(patient[2])
            self.middle_name_input.setText(patient[3])
            self.birth_date_input.setText(patient[4])
            self.gender_input.setCurrentText(patient[5])
            self.phone_input.setText(patient[6])
            self.email_input.setText(patient[7])
            self.address_input.setText(patient[8])
            self.passport_series_input.setText(patient[9])
            self.passport_number_input.setText(patient[10])
            self.insurance_number_input.setText(patient[11])

    def save_patient(self):
        first_name = self.first_name_input.text()
        last_name = self.last_name_input.text()
        middle_name = self.middle_name_input.text()
        birth_date = self.birth_date_input.text()
        gender = self.gender_input.currentText()
        phone = self.phone_input.text()
        email = self.email_input.text()
        address = self.address_input.text()
        passport_series = self.passport_series_input.text()
        passport_number = self.passport_number_input.text()
        insurance_number = self.insurance_number_input.text()

        if not Validator.validate_name(first_name):
            QMessageBox.warning(self, "Ошибка", "Имя должно начинаться с заглавной буквы и содержать только русские буквы")
            return

        if not Validator.validate_name(last_name):
            QMessageBox.warning(self, "Ошибка", "Фамилия должна начинаться с заглавной буквы и содержать только русские буквы")
            return

        if not Validator.validate_name(middle_name):
            QMessageBox.warning(self, "Ошибка", "Отчество должно начинаться с заглавной буквы и содержать только русские буквы")
            return

        if not Validator.validate_gender(gender):
            QMessageBox.warning(self, "Ошибка", "Пол должен быть 'М' или 'Ж'")
            return

        if not Validator.validate_phone(phone):
            QMessageBox.warning(self, "Ошибка", "Телефон должен быть в формате +7XXXXXXXXXX")
            return

        if not Validator.validate_address(address):
            QMessageBox.warning(self, "Ошибка", "Адрес должен содержать только русские буквы, цифры и знаки препинания")
            return

        if not Validator.validate_passport_series(passport_series):
            QMessageBox.warning(self, "Ошибка", "Серия паспорта должна состоять из 4 цифр")
            return

        if not Validator.validate_passport_number(passport_number):
            QMessageBox.warning(self, "Ошибка", "Номер паспорта должен состоять из 6 цифр")
            return

        if not Validator.validate_insurance_number(insurance_number):
            QMessageBox.warning(self, "Ошибка", "Номер страховки должен состоять из 16 цифр")
            return

        try:
            cursor.execute("""
                UPDATE Patient
                SET first_name=?, last_name=?, middle_name=?, birth_date=?, gender=?, phone=?, email=?, address=?, passport_series=?, passport_number=?, insurance_number=?
                WHERE patient_id=?
            """, (first_name, last_name, middle_name, birth_date, gender, phone, email, address, passport_series, passport_number, insurance_number, self.patient_id))
            conn.commit()
            QMessageBox.information(self, "Успех", "Данные пациента успешно обновлены")
            self.main_window.update_patient_table()  # Обновляем таблицу пациентов
            self.close()
        except sqlite3.Error as e:
            QMessageBox.warning(self, "Ошибка", f"Ошибка при обновлении данных пациента: {e}")

# Окно редактирования данных врача (только для админа)
class EditDoctorWindow(QWidget):
    def __init__(self, main_window, doctor_id):
        super().__init__()
        self.main_window = main_window
        self.doctor_id = doctor_id
        self.setWindowTitle("Редактирование данных врача")
        self.setGeometry(100, 100, 300, 250)

        layout = QVBoxLayout()

        self.first_name_input = QLineEdit(self)
        self.first_name_input.setPlaceholderText("Имя")
        layout.addWidget(self.first_name_input)

        self.last_name_input = QLineEdit(self)
        self.last_name_input.setPlaceholderText("Фамилия")
        layout.addWidget(self.last_name_input)

        self.middle_name_input = QLineEdit(self)
        self.middle_name_input.setPlaceholderText("Отчество")
        layout.addWidget(self.middle_name_input)

        self.specialization_input = QLineEdit(self)
        self.specialization_input.setPlaceholderText("Специализация")
        layout.addWidget(self.specialization_input)

        self.phone_input = QLineEdit(self)
        self.phone_input.setPlaceholderText("Телефон")
        layout.addWidget(self.phone_input)

        self.email_input = QLineEdit(self)
        self.email_input.setPlaceholderText("Email")
        layout.addWidget(self.email_input)

        self.save_button = QPushButton("Сохранить", self)
        self.save_button.clicked.connect(self.save_doctor)
        layout.addWidget(self.save_button)

        self.setLayout(layout)

        self.load_doctor_data()

    def load_doctor_data(self):
        cursor.execute("SELECT * FROM Doctor WHERE doctor_id=?", (self.doctor_id,))
        doctor = cursor.fetchone()
        if doctor:
            self.first_name_input.setText(doctor[1])
            self.last_name_input.setText(doctor[2])
            self.middle_name_input.setText(doctor[3])
            self.specialization_input.setText(doctor[4])
            self.phone_input.setText(doctor[5])
            self.email_input.setText(doctor[6])

    def save_doctor(self):
        first_name = self.first_name_input.text()
        last_name = self.last_name_input.text()
        middle_name = self.middle_name_input.text()
        specialization = self.specialization_input.text()
        phone = self.phone_input.text()
        email = self.email_input.text()

        try:
            cursor.execute("""
                UPDATE Doctor
                SET first_name=?, last_name=?, middle_name=?, specialization=?, phone=?, email=?
                WHERE doctor_id=?
            """, (first_name, last_name, middle_name, specialization, phone, email, self.doctor_id))
            conn.commit()
            QMessageBox.information(self, "Успех", "Данные врача успешно обновлены")
            self.main_window.update_doctor_table()  # Обновляем таблицу врачей
            self.close()
        except sqlite3.Error as e:
            QMessageBox.warning(self, "Ошибка", f"Ошибка при обновлении данных врача: {e}")

# Основное окно приложения
class MainWindow(QMainWindow):
    def __init__(self, role):
        super().__init__()
        self.role = role
        self.setWindowTitle("Стоматологическая клиника")
        self.setGeometry(100, 100, 1200, 600)

        # Создание главного виджета и компоновки
        main_widget = QWidget()
        self.setCentralWidget(main_widget)
        layout = QVBoxLayout(main_widget)

        # Создание вкладок
        self.tabs = QTabWidget()
        layout.addWidget(self.tabs)

        # Вкладка для пациентов
        self.patient_tab = QWidget()
        self.tabs.addTab(self.patient_tab, "Пациенты")
        patient_layout = QVBoxLayout(self.patient_tab)

        self.patient_table = QTableWidget(self)
        self.patient_table.setColumnCount(14)
        self.patient_table.setHorizontalHeaderLabels(["ID", "Имя", "Фамилия", "Отчество", "Дата рождения", "Пол", "Телефон", "Email", "Адрес", "Серия паспорта", "Номер паспорта", "Номер страховки", "Имя пользователя", "Пароль"])
        self.patient_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        patient_layout.addWidget(self.patient_table)

        self.add_patient_button = QPushButton("Добавить пациента", self)
        self.add_patient_button.clicked.connect(self.show_add_patient_window)
        patient_layout.addWidget(self.add_patient_button)

        # Добавляем кнопку удаления пациента
        self.delete_patient_button = QPushButton("Удалить пациента", self)
        self.delete_patient_button.clicked.connect(self.delete_patient)
        patient_layout.addWidget(self.delete_patient_button)

        # Добавляем кнопку редактирования пациента
        self.edit_patient_button = QPushButton("Редактировать пациента", self)
        self.edit_patient_button.clicked.connect(self.edit_patient)
        patient_layout.addWidget(self.edit_patient_button)

        # Добавляем кнопку экспорта медицинских карт в Excel
        self.export_medical_records_button = QPushButton("Экспорт медицинских карт в Excel", self)
        self.export_medical_records_button.clicked.connect(self.export_medical_records_to_excel)
        patient_layout.addWidget(self.export_medical_records_button)

        # Вкладка для врачей
        self.doctor_tab = QWidget()
        self.tabs.addTab(self.doctor_tab, "Врачи")
        doctor_layout = QVBoxLayout(self.doctor_tab)

        self.doctor_table = QTableWidget(self)
        self.doctor_table.setColumnCount(9)
        self.doctor_table.setHorizontalHeaderLabels(["ID Врача", "Имя", "Фамилия", "Отчество", "Специализация", "Телефон", "Email", "Имя пользователя", "Пароль"])
        self.doctor_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        doctor_layout.addWidget(self.doctor_table)

        self.register_doctor_button = QPushButton("Зарегистрировать врача", self)
        self.register_doctor_button.clicked.connect(self.show_register_doctor_window)
        doctor_layout.addWidget(self.register_doctor_button)

        # Добавляем кнопку удаления врача
        self.delete_doctor_button = QPushButton("Удалить врача", self)
        self.delete_doctor_button.clicked.connect(self.delete_doctor)
        doctor_layout.addWidget(self.delete_doctor_button)

        # Добавляем кнопку редактирования врача
        self.edit_doctor_button = QPushButton("Редактировать врача", self)
        self.edit_doctor_button.clicked.connect(self.edit_doctor)
        doctor_layout.addWidget(self.edit_doctor_button)

        # Вкладка для приемов
        self.appointment_tab = QWidget()
        self.tabs.addTab(self.appointment_tab, "Приемы")
        appointment_layout = QVBoxLayout(self.appointment_tab)

        self.appointment_table = QTableWidget(self)
        self.appointment_table.setColumnCount(6)
        self.appointment_table.setHorizontalHeaderLabels(["ID Приема", "Имя пациента", "Имя врача", "Дата приема", "Статус", "Специализация"])
        self.appointment_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        appointment_layout.addWidget(self.appointment_table)

        # Кнопка для подтверждения записи на прием
        self.confirm_appointment_button = QPushButton("Подтвердить запись", self)
        self.confirm_appointment_button.clicked.connect(self.confirm_appointment)
        appointment_layout.addWidget(self.confirm_appointment_button)

        # Кнопка для редактирования данных о приеме
        self.edit_appointment_button = QPushButton("Редактировать данные", self)
        self.edit_appointment_button.clicked.connect(self.edit_appointment)
        appointment_layout.addWidget(self.edit_appointment_button)

        # Добавляем кнопку удаления приема
        self.delete_appointment_button = QPushButton("Удалить прием", self)
        self.delete_appointment_button.clicked.connect(self.delete_appointment)
        appointment_layout.addWidget(self.delete_appointment_button)

        # Вкладка для расписания врача
        self.doctor_schedule_tab = QWidget()
        self.tabs.addTab(self.doctor_schedule_tab, "Расписание врача")
        doctor_schedule_layout = QVBoxLayout(self.doctor_schedule_tab)

        self.doctor_schedule_table = QTableWidget(self)
        self.doctor_schedule_table.setColumnCount(4)
        self.doctor_schedule_table.setHorizontalHeaderLabels(["ID Приема", "Имя пациента", "Дата приема", "Статус"])
        self.doctor_schedule_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        doctor_schedule_layout.addWidget(self.doctor_schedule_table)

        # Отображение данных при запуске программы
        self.update_patient_table()
        self.update_doctor_table()
        self.update_appointment_table()

        # Обработка события нажатия на строку в таблице врачей
        self.doctor_table.cellClicked.connect(self.show_doctor_schedule)

    def show_add_patient_window(self):
        self.add_patient_window = AddPatientWindow(self)
        self.add_patient_window.show()

    def show_register_doctor_window(self):
        self.register_doctor_window = DoctorRegisterWindow(self)
        self.register_doctor_window.show()

    def delete_patient(self):
        selected_row = self.patient_table.currentRow()
        if selected_row == -1:
            QMessageBox.warning(self, "Ошибка", "Выберите пациента для удаления")
            return

        patient_id = self.patient_table.item(selected_row, 0).text()

        try:
            cursor.execute("DELETE FROM Patient WHERE patient_id=?", (patient_id,))
            conn.commit()
            QMessageBox.information(self, "Успех", "Пациент успешно удален")
            self.update_patient_table()  # Обновляем таблицу пациентов
        except sqlite3.Error as e:
            QMessageBox.warning(self, "Ошибка", f"Ошибка при удалении пациента: {e}")

    def delete_doctor(self):
        selected_row = self.doctor_table.currentRow()
        if selected_row == -1:
            QMessageBox.warning(self, "Ошибка", "Выберите врача для удаления")
            return

        doctor_id = self.doctor_table.item(selected_row, 0).text()

        try:
            cursor.execute("DELETE FROM Doctor WHERE doctor_id=?", (doctor_id,))
            conn.commit()
            QMessageBox.information(self, "Успех", "Врач успешно удален")
            self.update_doctor_table()  # Обновляем таблицу врачей
        except sqlite3.Error as e:
            QMessageBox.warning(self, "Ошибка", f"Ошибка при удалении врача: {e}")

    def delete_appointment(self):
        selected_row = self.appointment_table.currentRow()
        if selected_row == -1:
            QMessageBox.warning(self, "Ошибка", "Выберите прием для удаления")
            return

        appointment_id = self.appointment_table.item(selected_row, 0).text()

        try:
            cursor.execute("DELETE FROM Appointment WHERE appointment_id=?", (appointment_id,))
            conn.commit()
            QMessageBox.information(self, "Успех", "Прием успешно удален")
            self.update_appointment_table()  # Обновляем таблицу приемов
        except sqlite3.Error as e:
            QMessageBox.warning(self, "Ошибка", f"Ошибка при удалении приема: {e}")

    def edit_patient(self):
        selected_row = self.patient_table.currentRow()
        if selected_row == -1:
            QMessageBox.warning(self, "Ошибка", "Выберите пациента для редактирования")
            return

        patient_id = self.patient_table.item(selected_row, 0).text()
        self.edit_patient_window = EditPatientWindow(self, patient_id)
        self.edit_patient_window.show()

    def edit_doctor(self):
        selected_row = self.doctor_table.currentRow()
        if selected_row == -1:
            QMessageBox.warning(self, "Ошибка", "Выберите врача для редактирования")
            return

        doctor_id = self.doctor_table.item(selected_row, 0).text()
        self.edit_doctor_window = EditDoctorWindow(self, doctor_id)
        self.edit_doctor_window.show()

    def confirm_appointment(self):
        selected_row = self.appointment_table.currentRow()
        if selected_row == -1:
            QMessageBox.warning(self, "Ошибка", "Выберите запись для подтверждения")
            return

        appointment_id = self.appointment_table.item(selected_row, 0).text()

        try:
            cursor.execute("UPDATE Appointment SET status='Запланировано' WHERE appointment_id=?", (appointment_id,))
            conn.commit()
            QMessageBox.information(self, "Успех", "Запись на прием подтверждена")
            self.update_appointment_table()  # Обновляем таблицу приемов
        except sqlite3.Error as e:
            QMessageBox.warning(self, "Ошибка", f"Ошибка при подтверждении записи на прием: {e}")

    def edit_appointment(self):
        selected_row = self.appointment_table.currentRow()
        if selected_row == -1:
            QMessageBox.warning(self, "Ошибка", "Выберите запись для редактирования")
            return

        appointment_id = self.appointment_table.item(selected_row, 0).text()
        self.edit_appointment_window = EditAppointmentWindow(self, appointment_id)
        self.edit_appointment_window.show()

    def update_patient_table(self):
        show_patients(self.patient_table)

    def update_doctor_table(self):
        show_doctors(self.doctor_table)

    def update_appointment_table(self):
        show_appointments(self.appointment_table)
        appointments = get_appointments()
        self.appointment_table.setRowCount(len(appointments))
        for row, appointment in enumerate(appointments):
            for col, item in enumerate(appointment):
                item = QTableWidgetItem(str(item))
                if col == 4:  # Статус находится в 5-ой колонке (индекс 4)
                    if item.text() == 'Запрошено':
                        item.setBackground(QColor(255, 0, 0))  # Красный цвет для запрошенных записей
                    elif item.text() == 'Запланировано':
                        item.setBackground(QColor(0, 255, 0))  # Зеленый цвет для запланированных записей
                    elif item.text() == 'Отменено':
                        item.setBackground(QColor(255, 255, 0))  # Желтый цвет для отмененных записей
                self.appointment_table.setItem(row, col, item)

    def show_doctor_schedule(self, row, col):
        doctor_id = self.doctor_table.item(row, 0).text()
        self.update_doctor_schedule_table(doctor_id)
        self.tabs.setCurrentWidget(self.doctor_schedule_tab)

    def update_doctor_schedule_table(self, doctor_id):
        cursor.execute("""
            SELECT a.appointment_id, p.first_name || ' ' || p.middle_name || ' ' || p.last_name AS patient_name, 
                   a.appointment_date, a.status
            FROM Appointment a
            JOIN Patient p ON a.patient_id = p.patient_id
            WHERE a.doctor_id = ?
            ORDER BY a.appointment_date
        """, (doctor_id,))
        appointments = cursor.fetchall()
        self.doctor_schedule_table.setRowCount(len(appointments))
        for row, appointment in enumerate(appointments):
            for col, item in enumerate(appointment):
                item = QTableWidgetItem(str(item))
                if appointment[3] == 'Запрошено':
                    item.setBackground(QColor(255, 0, 0))  # Красный цвет для запрошенных записей
                self.doctor_schedule_table.setItem(row, col, item)

    def export_medical_records_to_excel(self):
        # Создаем новую книгу Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Медицинские карты"

        # Заголовки столбцов
        headers = ["ID Лечения", "Имя пациента", "Имя врача", "Начало лечения", "Окончание лечения", "Диагноз",
                   "План лечения"]
        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            cell = ws[f"{col_letter}1"]
            cell.value = header
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                                 bottom=Side(style='thin'))
            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

        # Получаем данные из базы данных
        cursor.execute("""
            SELECT t.treatment_id, p.first_name || ' ' || p.last_name AS patient_name, 
                   d.first_name || ' ' || d.last_name AS doctor_name, t.start_date, t.end_date, t.diagnosis, t.treatment_plan
            FROM Treatment t
            JOIN Patient p ON t.patient_id = p.patient_id
            JOIN Doctor d ON t.doctor_id = d.doctor_id
        """)
        treatments = cursor.fetchall()

        # Заполняем данные в Excel
        for row_num, treatment in enumerate(treatments, 2):
            for col_num, item in enumerate(treatment, 1):
                col_letter = get_column_letter(col_num)
                cell = ws[f"{col_letter}{row_num}"]
                cell.value = item
                cell.alignment = Alignment(horizontal='left', vertical='center')
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                                     bottom=Side(style='thin'))

        # Автоматический размер столбцов
        for column_cells in ws.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = length + 2

        # Сохраняем файл
        wb.save("medical_records.xlsx")
        QMessageBox.information(self, "Успех", "Медицинские карты успешно экспортированы в файл medical_records.xlsx")

# Окно редактирования данных о приеме (только для админа)
class EditAppointmentWindow(QWidget):
    def __init__(self, main_window, appointment_id):
        super().__init__()
        self.main_window = main_window
        self.appointment_id = appointment_id
        self.setWindowTitle("Редактирование данных о приеме")
        self.setGeometry(100, 100, 300, 250)

        layout = QVBoxLayout()

        self.patient_name_input = QLineEdit(self)
        self.patient_name_input.setPlaceholderText("Имя пациента")
        layout.addWidget(self.patient_name_input)

        self.doctor_name_input = QLineEdit(self)
        self.doctor_name_input.setPlaceholderText("Имя врача")
        layout.addWidget(self.doctor_name_input)

        self.appointment_date_input = QLineEdit(self)
        self.appointment_date_input.setPlaceholderText("Дата приема")
        layout.addWidget(self.appointment_date_input)

        self.status_input = QLineEdit(self)
        self.status_input.setPlaceholderText("Статус")
        layout.addWidget(self.status_input)

        self.save_button = QPushButton("Сохранить", self)
        self.save_button.clicked.connect(self.save_appointment)
        layout.addWidget(self.save_button)

        self.setLayout(layout)

        self.load_appointment_data()

    def load_appointment_data(self):
        cursor.execute("""
            SELECT a.appointment_id, p.first_name || ' ' || p.middle_name || ' ' || p.last_name AS patient_name, 
                   d.first_name || ' ' || d.middle_name || ' ' || d.last_name AS doctor_name, a.appointment_date, a.status
            FROM Appointment a
            JOIN Patient p ON a.patient_id = p.patient_id
            JOIN Doctor d ON a.doctor_id = d.doctor_id
            WHERE a.appointment_id = ?
        """, (self.appointment_id,))
        appointment = cursor.fetchone()
        if appointment:
            self.patient_name_input.setText(appointment[1])
            self.doctor_name_input.setText(appointment[2])
            self.appointment_date_input.setText(appointment[3])
            self.status_input.setText(appointment[4])

    def save_appointment(self):
        patient_name = self.patient_name_input.text()
        doctor_name = self.doctor_name_input.text()
        appointment_date = self.appointment_date_input.text()
        status = self.status_input.text()

        try:
            cursor.execute("""
                UPDATE Appointment
                SET appointment_date=?, status=?
                WHERE appointment_id=?
            """, (appointment_date, status, self.appointment_id))
            conn.commit()
            QMessageBox.information(self, "Успех", "Данные о приеме успешно обновлены")
            self.main_window.update_appointment_table()  # Обновляем таблицу приемов
            self.close()
        except sqlite3.Error as e:
            QMessageBox.warning(self, "Ошибка", f"Ошибка при обновлении данных о приеме: {e}")

# Окно записи на прием
class AppointmentWindow(QWidget):
    def __init__(self, patient_id):
        super().__init__()
        self.patient_id = patient_id
        self.setWindowTitle("Запись на прием")
        self.setGeometry(100, 100, 300, 200)

        layout = QVBoxLayout()

        self.doctor_combo = QComboBox(self)
        cursor.execute("SELECT doctor_id, first_name, last_name FROM Doctor")
        doctors = cursor.fetchall()
        for doctor in doctors:
            self.doctor_combo.addItem(f"{doctor[1]} {doctor[2]}", doctor[0])
        layout.addWidget(self.doctor_combo)

        self.date_input = QLineEdit(self)
        self.date_input.setPlaceholderText("Дата приема (YYYY-MM-DD HH:MM)")
        layout.addWidget(self.date_input)

        self.request_button = QPushButton("Запросить запись", self)
        self.request_button.clicked.connect(self.request_appointment)
        layout.addWidget(self.request_button)

        self.setLayout(layout)

    def request_appointment(self):
        doctor_id = self.doctor_combo.currentData()
        appointment_date = self.date_input.text()

        try:
            cursor.execute("""
                INSERT INTO Appointment (patient_id, doctor_id, appointment_date, status, notes)
                VALUES (?, ?, ?, 'Запрошено', 'Жалоба на боль в зубе')
            """, (self.patient_id, doctor_id, appointment_date))
            conn.commit()
            QMessageBox.information(self, "Успех", f"Запись на прием к врачу на {appointment_date} успешно создана")
            self.close()
        except sqlite3.Error as e:
            QMessageBox.warning(self, "Ошибка", f"Ошибка при создании записи на прием: {e}")

# Окно пациента
class PatientWindow(QWidget):
    def __init__(self, patient):
        super().__init__()
        self.patient = patient
        self.setWindowTitle("Пациент")
        self.setGeometry(100, 100, 800, 600)

        layout = QVBoxLayout()

        self.welcome_label = QLabel(f"Добро пожаловать, {self.patient[1]} {self.patient[2]}", self)
        layout.addWidget(self.welcome_label)

        # Создание вкладок
        self.tabs = QTabWidget()
        layout.addWidget(self.tabs)

        # Вкладка для записей на прием
        self.appointments_tab = QWidget()
        self.tabs.addTab(self.appointments_tab, "Записи на прием")
        appointments_layout = QVBoxLayout(self.appointments_tab)

        self.appointments_table = QTableWidget(self)
        self.appointments_table.setColumnCount(5)
        self.appointments_table.setHorizontalHeaderLabels(["ID Приема", "Имя врача", "Дата приема", "Статус", "Примечания"])
        self.appointments_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        appointments_layout.addWidget(self.appointments_table)

        # Подключаем обработчик события нажатия на строку в таблице "Записи на прием"
        self.appointments_table.cellClicked.connect(self.on_appointment_clicked)

        # Вкладка для информации о врачах
        self.doctors_tab = QWidget()
        self.tabs.addTab(self.doctors_tab, "Врачи")
        doctors_layout = QVBoxLayout(self.doctors_tab)

        self.doctors_table = QTableWidget(self)
        self.doctors_table.setColumnCount(6)
        self.doctors_table.setHorizontalHeaderLabels(["ID Врача", "Имя", "Фамилия", "Специализация", "Телефон", "Email"])
        self.doctors_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        doctors_layout.addWidget(self.doctors_table)

        # Вкладка для медицинской карты
        self.medical_record_tab = QWidget()
        self.tabs.addTab(self.medical_record_tab, "Медицинская карта")
        medical_record_layout = QVBoxLayout(self.medical_record_tab)

        self.medical_record_table = QTableWidget(self)
        self.medical_record_table.setColumnCount(6)
        self.medical_record_table.setHorizontalHeaderLabels(["ID Лечения", "Имя врача", "Начало лечения", "Окончание лечения", "Диагноз", "План лечения"])
        self.medical_record_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        medical_record_layout.addWidget(self.medical_record_table)

        # Подключаем обработчик события нажатия на строку в таблице медицинской карты
        self.medical_record_table.cellClicked.connect(self.on_medical_record_clicked)

        # Кнопка для записи на прием
        self.appointment_button = QPushButton("Записаться на прием", self)
        self.appointment_button.clicked.connect(self.show_appointment_window)
        layout.addWidget(self.appointment_button)

        # Добавляем QLabel с информацией о номере телефона администратора внизу окна
        self.admin_phone_label = QLabel("По всем вопросам обращайтесь по телефону: +7 (123) 456-78-90", self)
        layout.addWidget(self.admin_phone_label)

        self.setLayout(layout)

        # Заполнение таблиц данными
        self.update_appointments_table()
        self.update_doctors_table()
        self.update_medical_record_table()

    def on_appointment_clicked(self, row, col):
        # Получаем appointment_id и doctor_id выбранной строки
        appointment_id = self.appointments_table.item(row, 0).text()
        doctor_name = self.appointments_table.item(row, 1).text()
        doctor_id = self.get_doctor_id_by_name(doctor_name)

        # Загружаем соответствующие записи из медицинской карты
        self.load_medical_records_for_appointment(appointment_id, doctor_id)
        # Переключаемся на вкладку "Медицинская карта"
        self.tabs.setCurrentWidget(self.medical_record_tab)

    def on_medical_record_clicked(self, row, col):
        # Получаем данные выбранной строки
        treatment_id = self.medical_record_table.item(row, 0).text()
        doctor_name = self.medical_record_table.item(row, 1).text()
        start_date = self.medical_record_table.item(row, 2).text()
        end_date = self.medical_record_table.item(row, 3).text()
        diagnosis = self.medical_record_table.item(row, 4).text()
        treatment_plan = self.medical_record_table.item(row, 5).text()

        # Выводим информацию в консоль
        print(f"Treatment ID: {treatment_id}")
        print(f"Doctor Name: {doctor_name}")
        print(f"Start Date: {start_date}")
        print(f"End Date: {end_date}")
        print(f"Diagnosis: {diagnosis}")
        print(f"Treatment Plan: {treatment_plan}")

    def update_appointments_table(self):
        cursor.execute("""
            SELECT a.appointment_id, d.first_name || ' ' || d.last_name AS doctor_name, a.appointment_date, a.status, a.notes
            FROM Appointment a
            JOIN Doctor d ON a.doctor_id = d.doctor_id
            WHERE a.patient_id = ?
        """, (self.patient[0],))
        appointments = cursor.fetchall()
        self.appointments_table.setRowCount(len(appointments))
        for row, appointment in enumerate(appointments):
            for col, item in enumerate(appointment):
                self.appointments_table.setItem(row, col, QTableWidgetItem(str(item)))

    def update_doctors_table(self):
        cursor.execute("SELECT doctor_id, first_name, last_name, specialization, phone, email FROM Doctor")
        doctors = cursor.fetchall()
        self.doctors_table.setRowCount(len(doctors))
        for row, doctor in enumerate(doctors):
            for col, item in enumerate(doctor):
                self.doctors_table.setItem(row, col, QTableWidgetItem(str(item)))

    def update_medical_record_table(self):
        cursor.execute("""
            SELECT t.treatment_id, d.first_name || ' ' || d.last_name AS doctor_name,
             t.start_date, t.end_date, t.diagnosis, t.treatment_plan
            FROM Treatment t
            JOIN Doctor d ON t.doctor_id = d.doctor_id
            WHERE t.patient_id = ?
        """, (self.patient[0],))
        treatments = cursor.fetchall()
        self.medical_record_table.setRowCount(len(treatments))
        for row, treatment in enumerate(treatments):
            for col, item in enumerate(treatment):
                self.medical_record_table.setItem(row, col, QTableWidgetItem(str(item)))

    def load_medical_records_for_appointment(self, appointment_id, doctor_id):
        cursor.execute("""
            SELECT t.treatment_id, d.first_name || ' ' || d.last_name AS doctor_name, t.start_date, t.end_date, t.diagnosis, t.treatment_plan
            FROM Treatment t
            JOIN Doctor d ON t.doctor_id = d.doctor_id
            JOIN Appointment a ON t.patient_id = a.patient_id
            WHERE a.appointment_id = ? AND d.doctor_id = ?
        """, (appointment_id, doctor_id))
        treatments = cursor.fetchall()
        self.medical_record_table.setRowCount(len(treatments))
        for row, treatment in enumerate(treatments):
            for col, item in enumerate(treatment):
                self.medical_record_table.setItem(row, col, QTableWidgetItem(str(item)))

    def get_doctor_id_by_name(self, doctor_name):
        cursor.execute("""
            SELECT doctor_id
            FROM Doctor
            WHERE first_name || ' ' || last_name = ?
        """, (doctor_name,))
        result = cursor.fetchone()
        if result:
            return result[0]
        return None

    def show_appointment_window(self):
        self.appointment_window = AppointmentWindow(self.patient[0])
        self.appointment_window.show()

# Окно врача
class DoctorWindow(QMainWindow):
    def __init__(self, doctor):
        super().__init__()
        self.doctor = doctor
        self.setWindowTitle(f"Врач - {self.doctor[1]} {self.doctor[2]}")
        self.setGeometry(100, 100, 800, 600)

        # Central widget and layout
        main_widget = QWidget()
        self.setCentralWidget(main_widget)
        layout = QVBoxLayout(main_widget)

        # Welcome label
        self.welcome_label = QLabel(f"Добро пожаловать, {self.doctor[1]} {self.doctor[2]}", self)
        layout.addWidget(self.welcome_label)

        # Tabs for different sections
        self.tabs = QTabWidget()
        layout.addWidget(self.tabs)

        # Appointments tab
        self.appointments_tab = QWidget()
        self.tabs.addTab(self.appointments_tab, "Мои Приемы")
        appointments_layout = QVBoxLayout(self.appointments_tab)

        self.appointments_table = QTableWidget(self)
        self.appointments_table.setColumnCount(5)
        self.appointments_table.setHorizontalHeaderLabels(["ID Приема", "Имя пациента", "Дата", "Статус", "Примечания"])
        self.appointments_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        appointments_layout.addWidget(self.appointments_table)

        self.appointments_table.cellClicked.connect(self.on_appointment_selected)

        self.edit_appointment_button = QPushButton("Редактировать прием", self)
        self.edit_appointment_button.clicked.connect(self.edit_appointment)
        appointments_layout.addWidget(self.edit_appointment_button)

        # Patient Records tab
        self.patient_records_tab = QWidget()
        self.tabs.addTab(self.patient_records_tab, "Медицинские Записи")
        self.patient_records_layout = QVBoxLayout(self.patient_records_tab)

        self.patient_records_table = QTableWidget(self)
        self.patient_records_table.setColumnCount(6)
        self.patient_records_table.setHorizontalHeaderLabels(["ID Лечения", "Начало лечения", "Окончание лечения", "Диагноз", "План лечения", "Действия"])
        self.patient_records_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.patient_records_layout.addWidget(self.patient_records_table)

        self.add_treatment_button = QPushButton("Добавить запись", self)
        self.add_treatment_button.clicked.connect(self.add_treatment)
        self.patient_records_layout.addWidget(self.add_treatment_button)

        # Load data when the window is opened
        self.load_appointments()

    def load_appointments(self):
        cursor.execute("""
            SELECT a.appointment_id, p.first_name || ' ' || p.last_name AS patient_name, 
                   a.appointment_date, a.status, a.notes
            FROM Appointment a
            JOIN Patient p ON a.patient_id = p.patient_id
            WHERE a.doctor_id = ?
        """, (self.doctor[0],))
        appointments = cursor.fetchall()
        self.appointments_table.setRowCount(len(appointments))
        for row, appointment in enumerate(appointments):
            for col, item in enumerate(appointment):
                self.appointments_table.setItem(row, col, QTableWidgetItem(str(item)))

    def on_appointment_selected(self, row, column):
        appointment_id = self.appointments_table.item(row, 0).text()
        # Fetch patient_id from appointment_id
        cursor.execute("SELECT patient_id FROM Appointment WHERE appointment_id = ?", (appointment_id,))
        patient_id_result = cursor.fetchone()
        if patient_id_result:
            self.patient_id = patient_id_result[0]
            self.load_patient_medical_records(self.patient_id)
            self.tabs.setCurrentWidget(self.patient_records_tab)

    def load_patient_medical_records(self, patient_id):
        cursor.execute("""
            SELECT t.treatment_id, t.start_date, t.end_date, t.diagnosis, t.treatment_plan
            FROM Treatment t
            WHERE t.patient_id = ?
        """, (patient_id,))
        treatments = cursor.fetchall()
        self.patient_records_table.setRowCount(len(treatments))
        for row, treatment in enumerate(treatments):
            for col, item in enumerate(treatment):
                self.patient_records_table.setItem(row, col, QTableWidgetItem(str(item)))
            # Add an "Edit" button in the "Actions" column
            edit_button = QPushButton("Редактировать", self)
            edit_button.clicked.connect(lambda _, tid=treatment[0]: self.edit_treatment(tid))
            self.patient_records_table.setCellWidget(row, 5, edit_button)

    def add_treatment(self):
        selected_row = self.appointments_table.currentRow()
        if selected_row == -1:
            QMessageBox.warning(self, "Ошибка", "Выберите пациента для добавления медицинской записи")
            return

        appointment_id = self.appointments_table.item(selected_row, 0).text()
        # Fetch patient_id from appointment_id
        cursor.execute("SELECT patient_id FROM Appointment WHERE appointment_id = ?", (appointment_id,))
        patient_id_result = cursor.fetchone()
        if patient_id_result:
            patient_id = patient_id_result[0]
            self.add_treatment_window = AddTreatmentWindow(self, patient_id)
            self.add_treatment_window.show()

    def edit_treatment(self, treatment_id):
        self.edit_treatment_window = EditTreatmentWindow(self, treatment_id)
        self.edit_treatment_window.show()

    def edit_appointment(self):
        selected_row = self.appointments_table.currentRow()
        if selected_row == -1:
            QMessageBox.warning(self, "Ошибка", "Выберите прием для редактирования")
            return

        appointment_id = self.appointments_table.item(selected_row, 0).text()
        self.edit_appointment_window = EditDoctorAppointmentWindow(self, appointment_id)
        self.edit_appointment_window.show()

# Окно добавления медицинской записи
class AddTreatmentWindow(QWidget):
    def __init__(self, doctor_window, patient_id):
        super().__init__()
        self.doctor_window = doctor_window
        self.patient_id = patient_id
        self.setWindowTitle("Добавить медицинскую запись")
        self.setGeometry(100, 100, 300, 300)

        layout = QVBoxLayout()

        self.start_date_label = QLabel("Начало лечения:", self)
        layout.addWidget(self.start_date_label)

        self.start_date_input = QLineEdit(self)
        layout.addWidget(self.start_date_input)

        self.end_date_label = QLabel("Окончание лечения:", self)
        layout.addWidget(self.end_date_label)

        self.end_date_input = QLineEdit(self)
        layout.addWidget(self.end_date_input)

        self.diagnosis_label = QLabel("Диагноз:", self)
        layout.addWidget(self.diagnosis_label)

        self.diagnosis_input = QLineEdit(self)
        layout.addWidget(self.diagnosis_input)

        self.treatment_plan_label = QLabel("План лечения:", self)
        layout.addWidget(self.treatment_plan_label)

        self.treatment_plan_input = QTextEdit(self)
        layout.addWidget(self.treatment_plan_input)

        self.save_button = QPushButton("Сохранить", self)
        self.save_button.clicked.connect(self.save_treatment)
        layout.addWidget(self.save_button)

        self.setLayout(layout)

    def save_treatment(self):
        start_date = self.start_date_input.text()
        end_date = self.end_date_input.text()
        diagnosis = self.diagnosis_input.text()
        treatment_plan = self.treatment_plan_input.toPlainText()

        try:
            cursor.execute("""
                INSERT INTO Treatment (patient_id, doctor_id, start_date, end_date, diagnosis, treatment_plan)
                VALUES (?, ?, ?, ?, ?, ?)
            """, (self.patient_id, self.doctor_window.doctor[0], start_date, end_date, diagnosis, treatment_plan))
            conn.commit()
            QMessageBox.information(self, "Успех", "Медицинская запись успешно добавлена")
            self.doctor_window.load_patient_medical_records(self.patient_id)
            self.close()
        except sqlite3.Error as e:
            QMessageBox.warning(self, "Ошибка", f"Ошибка при добавлении медицинской записи: {e}")

# Окно редактирования медицинской записи
class EditTreatmentWindow(QWidget):
    def __init__(self, doctor_window, treatment_id):
        super().__init__()
        self.doctor_window = doctor_window
        self.treatment_id = treatment_id
        self.setWindowTitle("Редактирование медицинской записи")
        self.setGeometry(100, 100, 300, 300)

        layout = QVBoxLayout()

        self.start_date_label = QLabel("Начало лечения:", self)
        layout.addWidget(self.start_date_label)

        self.start_date_input = QLineEdit(self)
        layout.addWidget(self.start_date_input)

        self.end_date_label = QLabel("Окончание лечения:", self)
        layout.addWidget(self.end_date_label)

        self.end_date_input = QLineEdit(self)
        layout.addWidget(self.end_date_input)

        self.diagnosis_label = QLabel("Диагноз:", self)
        layout.addWidget(self.diagnosis_label)

        self.diagnosis_input = QLineEdit(self)
        layout.addWidget(self.diagnosis_input)

        self.treatment_plan_label = QLabel("План лечения:", self)
        layout.addWidget(self.treatment_plan_label)

        self.treatment_plan_input = QTextEdit(self)
        layout.addWidget(self.treatment_plan_input)

        self.save_button = QPushButton("Сохранить", self)
        self.save_button.clicked.connect(self.save_treatment)
        layout.addWidget(self.save_button)

        self.setLayout(layout)

        self.load_treatment_data()

    def load_treatment_data(self):
        cursor.execute("""
            SELECT start_date, end_date, diagnosis, treatment_plan
            FROM Treatment
            WHERE treatment_id = ?
        """, (self.treatment_id,))
        treatment = cursor.fetchone()
        if treatment:
            self.start_date_input.setText(treatment[0])
            self.end_date_input.setText(treatment[1])
            self.diagnosis_input.setText(treatment[2])
            self.treatment_plan_input.setPlainText(treatment[3])

    def save_treatment(self):
        start_date = self.start_date_input.text()
        end_date = self.end_date_input.text()
        diagnosis = self.diagnosis_input.text()
        treatment_plan = self.treatment_plan_input.toPlainText()

        try:
            cursor.execute("""
                UPDATE Treatment
                SET start_date=?, end_date=?, diagnosis=?, treatment_plan=?
                WHERE treatment_id=?
            """, (start_date, end_date, diagnosis, treatment_plan, self.treatment_id))
            conn.commit()
            QMessageBox.information(self, "Успех", "Медицинская запись успешно обновлена")
            self.doctor_window.load_patient_medical_records(self.doctor_window.patient_id)  # Обновляем таблицу медицинских записей
            self.close()
        except sqlite3.Error as e:
            QMessageBox.warning(self, "Ошибка", f"Ошибка при обновлении медицинской записи: {e}")

# Окно редактирования данных о приеме (только для врача)
class EditDoctorAppointmentWindow(QWidget):
    def __init__(self, doctor_window, appointment_id):
        super().__init__()
        self.doctor_window = doctor_window
        self.appointment_id = appointment_id
        self.setWindowTitle("Редактирование данных о приеме")
        self.setGeometry(100, 100, 300, 250)

        layout = QVBoxLayout()

        self.patient_name_input = QLineEdit(self)
        self.patient_name_input.setPlaceholderText("Имя пациента")
        layout.addWidget(self.patient_name_input)

        self.doctor_name_input = QLineEdit(self)
        self.doctor_name_input.setPlaceholderText("Имя врача")
        layout.addWidget(self.doctor_name_input)

        self.appointment_date_input = QLineEdit(self)
        self.appointment_date_input.setPlaceholderText("Дата приема")
        layout.addWidget(self.appointment_date_input)

        self.status_input = QLineEdit(self)
        self.status_input.setPlaceholderText("Статус")
        layout.addWidget(self.status_input)

        self.save_button = QPushButton("Сохранить", self)
        self.save_button.clicked.connect(self.save_appointment)
        layout.addWidget(self.save_button)

        self.setLayout(layout)

        self.load_appointment_data()

    def load_appointment_data(self):
        cursor.execute("""
            SELECT a.appointment_id, p.first_name || ' ' || p.middle_name || ' ' || p.last_name AS patient_name, 
                   d.first_name || ' ' || d.middle_name || ' ' || d.last_name AS doctor_name, a.appointment_date, a.status
            FROM Appointment a
            JOIN Patient p ON a.patient_id = p.patient_id
            JOIN Doctor d ON a.doctor_id = d.doctor_id
            WHERE a.appointment_id = ?
        """, (self.appointment_id,))
        appointment = cursor.fetchone()
        if appointment:
            self.patient_name_input.setText(appointment[1])
            self.doctor_name_input.setText(appointment[2])
            self.appointment_date_input.setText(appointment[3])
            self.status_input.setText(appointment[4])

    def save_appointment(self):
        patient_name = self.patient_name_input.text()
        doctor_name = self.doctor_name_input.text()
        appointment_date = self.appointment_date_input.text()
        status = self.status_input.text()

        try:
            cursor.execute("""
                UPDATE Appointment
                SET appointment_date=?, status=?
                WHERE appointment_id=?
            """, (appointment_date, status, self.appointment_id))
            conn.commit()
            QMessageBox.information(self, "Успех", "Данные о приеме успешно обновлены")
            self.doctor_window.load_appointments()  # Обновляем таблицу приемов
            self.close()
        except sqlite3.Error as e:
            QMessageBox.warning(self, "Ошибка", f"Ошибка при обновлении данных о приеме: {e}")

# Ensure database connection is open throughout the application
if __name__ == "__main__":
    app = QApplication(sys.argv)
    login_window = LoginWindow()
    login_window.show()
    sys.exit(app.exec())

# Close the database connection when the application exits
conn.close()